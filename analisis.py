import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Cargar los datos de ventas desde un archivo CSV o una base de datos
df_ventas = pd.read_csv('datos_ventas.csv')

# Crear un nuevo archivo de Excel
wb = Workbook()
ws = wb.active

# 1. Análisis de ventas por producto
ventas_por_producto = df_ventas.groupby('Producto')['Cantidad'].sum()
productos_populares = ventas_por_producto.nlargest(5)  # Obtener los 5 productos más populares
ws['A1'] = 'Producto'
ws['B1'] = 'Ventas'
for i, (producto, ventas) in enumerate(productos_populares.iteritems(), start=2):
    ws.cell(row=i, column=1, value=producto)
    ws.cell(row=i, column=2, value=ventas)

# 2. Análisis de ventas por canal
ventas_por_canal = df_ventas.groupby('Canal')['Monto'].sum()
ws = wb.create_sheet(title='Ventas por Canal')
for i, (canal, ventas) in enumerate(ventas_por_canal.iteritems(), start=1):
    ws.cell(row=i, column=1, value=canal)
    ws.cell(row=i, column=2, value=ventas)

# 3. Análisis de tendencias de ventas
df_ventas['Fecha'] = pd.to_datetime(df_ventas['Fecha'])  # Convertir la columna de fecha a formato datetime
df_ventas.set_index('Fecha', inplace=True)  # Establecer la fecha como el índice del DataFrame
ventas_mensuales = df_ventas.resample('M')['Monto'].sum()  # Resampleo de ventas por mes
ws = wb.create_sheet(title='Tendencias de Ventas Mensuales')
for i, (fecha, ventas) in enumerate(ventas_mensuales.iteritems(), start=1):
    ws.cell(row=i, column=1, value=fecha)
    ws.cell(row=i, column=2, value=ventas)

# 4. Análisis de clientes
ventas_por_cliente = df_ventas.groupby('Cliente')['Monto'].sum()
segmentos_clientes = pd.qcut(ventas_por_cliente, q=3, labels=['Bajo', 'Medio', 'Alto'])  # Segmentar clientes en 3 grupos
ws = wb.create_sheet(title='Segmentos de Clientes')
for i, (cliente, segmento) in enumerate(segmentos_clientes.iteritems(), start=1):
    ws.cell(row=i, column=1, value=cliente)
    ws.cell(row=i, column=2, value=segmento)

# 5. Análisis de efectividad de promociones
df_promociones = df_ventas[df_ventas['Promocion'] == True]
ventas_promociones = df_promociones.groupby('Promocion')['Monto'].sum()
ventas_totales = df_ventas['Monto'].sum()
porcentaje_ventas_promociones = ventas_promociones / ventas_totales * 100
ws = wb.create_sheet(title='Efectividad de Promociones')
ws['A1'] = 'Promoción'
ws['B1'] = 'Porcentaje de Ventas'
for i, (promocion, porcentaje) in enumerate(porcentaje_ventas_promociones.iteritems(), start=2):
    ws.cell(row=i, column=1, value=promocion)
    ws.cell(row=i, column=2, value=porcentaje)

# 6. Análisis de ventas por ubicación geográfica
ventas_por_ubicacion = df_ventas.groupby('Ubicacion')['Monto'].sum()
ws = wb.create_sheet(title='Ventas por Ubicación')
for i, (ubicacion, ventas) in enumerate(ventas_por_ubicacion.iteritems(), start=1):
    ws.cell(row=i, column=1, value=ubicacion)
    ws.cell(row=i, column=2, value=ventas)

# Agregar gráficos a las hojas de Excel
charts_ws = wb.create_sheet(title='Gráficos')
charts_ws.column_dimensions['A'].width = 30
charts_ws.column_dimensions['B'].width = 15

chart1 = plt.bar(productos_populares.index, productos_populares.values)
plt.xlabel('Producto')
plt.ylabel('Ventas')
plt.title('Productos Populares')
plt.xticks(rotation=90)
chart1_filename = 'chart1.png'
plt.savefig(chart1_filename)
plt.close()

img = openpyxl.drawing.image.Image(chart1_filename)
img.anchor = 'A1'
charts_ws.add_image(img)

chart2 = ventas_por_canal.plot(kind='bar')
plt.xlabel('Canal')
plt.ylabel('Ventas')
plt.title('Ventas por Canal')
plt.xticks(rotation=0)
chart2_filename = 'chart2.png'
plt.savefig(chart2_filename)
plt.close()

img = openpyxl.drawing.image.Image(chart2_filename)
img.anchor = 'A1'
charts_ws.add_image(img)

# Guardar el archivo Excel
wb.save('analisis_ventas.xlsx')